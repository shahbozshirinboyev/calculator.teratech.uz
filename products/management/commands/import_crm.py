from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path

import openpyxl
from django.core.management.base import BaseCommand

from products.models import CPU, RAM, MemoryType, MonoblockBase, Storage, StorageInterface


class Command(BaseCommand):
    help = "Import product data from data/crm.xlsx."

    def add_arguments(self, parser):
        parser.add_argument(
            "path",
            nargs="?",
            default="data/crm.xlsx",
            help="Excel file path. Default: data/crm.xlsx",
        )

    def handle(self, *args, **options):
        workbook_path = Path(options["path"])
        if not workbook_path.is_absolute():
            workbook_path = Path.cwd() / workbook_path
        if not workbook_path.exists():
            self.stderr.write(self.style.ERROR(f"File not found: {workbook_path}"))
            return

        workbook = openpyxl.load_workbook(workbook_path, data_only=True, read_only=True)
        total = 0

        for sheet in workbook.worksheets:
            total += self.import_sheet(sheet)

        self.stdout.write(self.style.SUCCESS(f"Imported {total} products from {workbook_path}"))

    def sheet_specs(self, name):
        normalized = name.strip().upper()
        if normalized == "H610":
            ram_type = MemoryType.DDR4
        elif normalized == "H61":
            ram_type = MemoryType.DDR3
        else:
            ram_type = MemoryType.DDR4
        return normalized, ram_type

    def import_sheet(self, sheet):
        chipset, ram_type = self.sheet_specs(sheet.title)
        bases = []
        imported = 0
        for column in range(2, sheet.max_column + 1, 2):
            header = str(sheet.cell(2, column).value or "").strip().upper()
            category = self.category_for(column, header)
            if not category:
                continue

            for row in range(3, sheet.max_row + 1):
                name = sheet.cell(row, column).value
                price = sheet.cell(row, column + 1).value
                if name is None or price is None:
                    continue

                name = str(name).strip()
                price = self.money(price)
                if category == "monitor":
                    base = self.upsert_base(name, price, chipset, ram_type)
                    bases.append(base)
                elif category == "cpu":
                    item, _ = CPU.objects.update_or_create(
                        name=name,
                        defaults={"price": price, "is_active": True},
                    )
                    item.compatible_bases.add(*bases)
                elif category == "ram":
                    RAM.objects.update_or_create(
                        name=name,
                        ram_type=ram_type,
                        defaults={"price": price, "capacity_gb": self.capacity_or_size(name), "is_active": True},
                    )
                elif category in {"ssd", "hdd"}:
                    Storage.objects.update_or_create(
                        name=name,
                        kind=category,
                        interface=self.storage_interface(name),
                        defaults={"price": price, "capacity_gb": self.storage_capacity(name), "is_active": True},
                    )
                imported += 1
        return imported

    def upsert_base(self, name, price, chipset, ram_type):
        size = self.capacity_or_size(name) or 0
        display_name = self.base_name(size, chipset, ram_type)
        base, _ = MonoblockBase.objects.update_or_create(
            name=display_name,
            motherboard_type=chipset,
            ram_type=ram_type,
            defaults={
                "price": price,
                "supports_nvme": chipset == "H610",
                "ram_slots": 2,
                "sata_ports": 1,
                "is_active": True,
            },
        )
        return base

    def base_name(self, size, chipset, ram_type):
        return f'{size}" FLAT IPS {chipset} {ram_type.upper()}'

    def category_for(self, column, header):
        if column == 2:
            return "monitor"
        if header == "CPU":
            return "cpu"
        if header in {"DDR3", "DDR4", "RAM"}:
            return "ram"
        if "SSD" in header:
            return "ssd"
        if "HDD" in header:
            return "hdd"
        return None

    def storage_interface(self, name):
        upper = name.upper()
        if "NVME" in upper or "M.2" in upper:
            return StorageInterface.NVME
        return StorageInterface.SATA

    def capacity_or_size(self, value):
        import re

        match = re.search(r"(\d+)", value)
        return int(match.group(1)) if match else None

    def storage_capacity(self, value):
        import re

        match = re.search(r"(\d+)\s*(TB|GB)", value.upper())
        if not match:
            return None
        amount = int(match.group(1))
        return amount * 1024 if match.group(2) == "TB" else amount

    def money(self, value):
        return Decimal(str(value)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
