import json
from decimal import Decimal, InvalidOperation

from django.contrib.auth import get_user_model
from django.contrib.auth.decorators import login_required
from django.db.models import Count, Q, Sum
from django.core.exceptions import PermissionDenied
from django.http import JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from django.views.decorators.http import require_POST

from .models import Order, OrderItem

User = get_user_model()

def _can_user_view_all_orders(user):
    return user.is_superuser or user.has_perm("orders.view_order") or user.has_perm("orders.change_order") or user.has_perm("orders.delete_order")


def _can_user_edit_order(user, order):
    return user.is_superuser or user.has_perm("orders.change_order") or order.sold_by_id == user.id


def _can_user_delete_order(user):
    return user.is_superuser or user.has_perm("orders.delete_order")


def _can_user_cancel_order(user, order):
    return _can_user_set_production_status(
        user,
        Order.ProductionStatus.CANCELLED,
        order.production_status,
        order,
    )


def _parse_decimal(value, default=Decimal("0")):
    try:
        return Decimal(str(value).replace(",", ".").strip())
    except (InvalidOperation, TypeError):
        return default


def _orders_for_user(user):
    qs = Order.objects.all()
    if _can_user_view_all_orders(user):
        return qs
    return qs.filter(sold_by=user)


def _serialize_choices(choices):
    return [{"value": value, "label": label} for value, label in choices]


def _can_user_change_production_status(user, current_status=None, order=None):
    if user.is_superuser or user.has_perm("orders.change_order"):
        return True
    if order is not None and order.sold_by_id != user.id:
        return False
    if not current_status:
        return True
    return current_status in {
        Order.ProductionStatus.AGREED,
        Order.ProductionStatus.QUEUED,
    }


def _allowed_production_status_choices_for_user(user, delivery_type, current_status=None, order=None):
    choices = Order.production_status_choices_for_delivery(delivery_type)
    
    # Agar o'zgartira olmasa - bo'sh
    if not _can_user_change_production_status(user, current_status, order):
        return []
    
    # Admin - barcha statuslarni ko'radi, lekin faqat oldinga va CANCELLED
    if user.is_superuser or user.has_perm("orders.change_order"):
        if current_status:
            status_flow = Order.production_status_flow(delivery_type)
            if current_status in status_flow:
                current_index = status_flow.index(current_status)
                # Faqat hozirgi va keyingi statuslar + CANCELLED
                filtered_choices = []
                for value, label in choices:
                    if value == Order.ProductionStatus.CANCELLED:
                        # CANCELLED har doim ko'rinadi
                        filtered_choices.append((value, label))
                    elif value in status_flow:
                        status_index = status_flow.index(value)
                        # Faqat hozirgi yoki keyingi statuslar
                        if status_index >= current_index:
                            filtered_choices.append((value, label))
                return filtered_choices
        return choices
    
    # Oddiy foydalanuvchi - faqat AGREED va QUEUED orasida
    allowed_statuses = {
        Order.ProductionStatus.AGREED,
        Order.ProductionStatus.QUEUED,
    }
    
    # Faqat oldinga o'tish uchun statuslarni filter qilish
    if current_status:
        status_flow = Order.production_status_flow(delivery_type)
        if current_status in status_flow:
            current_index = status_flow.index(current_status)
            # Faqat hozirgi va keyingi statuslar + CANCELLED
            filtered_choices = []
            for value, label in choices:
                if value == Order.ProductionStatus.CANCELLED:
                    # CANCELLED har doim ko'rinadi
                    filtered_choices.append((value, label))
                elif value in status_flow:
                    status_index = status_flow.index(value)
                    if status_index >= current_index and value in allowed_statuses:
                        filtered_choices.append((value, label))
                elif value in allowed_statuses:
                    filtered_choices.append((value, label))
            return filtered_choices
    
    return [(value, label) for value, label in choices if value in allowed_statuses]


def _all_production_status_choices_map(delivery_type):
    return dict(Order.production_status_choices_for_delivery(delivery_type))


def _can_user_set_production_status(user, target_status, current_status=None, order=None):
    if user.is_superuser or user.has_perm("orders.change_order"):
        return True
    if order is not None and order.sold_by_id != user.id:
        return False
    allowed_statuses = {
        Order.ProductionStatus.AGREED,
        Order.ProductionStatus.QUEUED,
    }
    if not current_status:
        return target_status in allowed_statuses
    if current_status in allowed_statuses:
        return target_status in allowed_statuses
    return bool(current_status) and target_status == current_status


def _validate_status_transition(order, target_status):
    """
    Status o'zgarishini tekshiradi. A variant: Faqat oldinga o'tish, to'lov tekshiruvi.
    Returns: (is_valid, error_message)
    """
    current_status = order.production_status
    payment_status = order.payment_status
    
    # 1. Kelishuvda -> Navbatda: To'lov holati tekshiruvi
    if current_status == Order.ProductionStatus.AGREED and target_status == Order.ProductionStatus.QUEUED:
        if payment_status == Order.PaymentStatus.UNPAID:
            return False, "To'lov holati 'To'lanmagan' bo'lsa, statusni 'Navbatda'ga o'tkazib bo'lmaydi"
    
    # 2. Yetkazildi statusiga har qanday to'lov holati bilan o'tish mumkin
    # (To'lanmagan, Yetkazilganda to'lanadi, Qisman to'langan, To'langan)
    
    # 3. CANCELLED (Bekor qilindi) - har qanday statusdan o'tish mumkin
    if target_status == Order.ProductionStatus.CANCELLED:
        return True, None
    
    # 4. Oldinga o'tish tekshiruvi - orqaga qaytish yo'q
    status_flow = Order.production_status_flow(order.delivery_type)
    if current_status in status_flow and target_status in status_flow:
        current_index = status_flow.index(current_status)
        target_index = status_flow.index(target_status)
        if target_index < current_index:
            return False, "Statusni orqaga qaytarib bo'lmaydi"
    
    return True, None


def _production_status_groups_json_for_user(user):
    groups = {}
    for delivery_type, _ in Order.DeliveryType.choices:
        groups[delivery_type] = _serialize_choices(
            _allowed_production_status_choices_for_user(user, delivery_type)
        )
    return json.dumps(groups, ensure_ascii=False)


def _all_production_status_groups_json():
    groups = {}
    for delivery_type, _ in Order.DeliveryType.choices:
        groups[delivery_type] = _serialize_choices(
            Order.production_status_choices_for_delivery(delivery_type)
        )
    return json.dumps(groups, ensure_ascii=False)


def _attach_order_permissions(user, orders):
    for order in orders:
        order.card_can_edit = _can_user_edit_order(user, order)
        order.card_can_delete = _can_user_delete_order(user)
        order.card_can_change_status = _can_user_change_production_status(
            user,
            order.production_status,
            order,
        )
        order.card_status_choices = _allowed_production_status_choices_for_user(
            user,
            order.delivery_type,
            order.production_status,
            order,
        )
    return orders


def _format_filter_date(value):
    if not value:
        return ""
    try:
        from datetime import datetime
        return datetime.strptime(value, "%Y-%m-%d").strftime("%d.%m.%Y")
    except (ValueError, TypeError):
        return value


def _order_status_tabs():
    return [
        {"value": "ALL", "icon": "📦", "label": "Barchasi", "type": "main"},
        {"type": "divider", "label": "Faol"},
        {"value": Order.ProductionStatus.AGREED, "icon": "🤝", "label": "Kelishuvda", "type": "status"},
        {"value": Order.ProductionStatus.QUEUED, "icon": "📋", "label": "Navbatda", "type": "status"},
        {"value": Order.ProductionStatus.IN_PROGRESS, "icon": "🛠", "label": "Tayyorlanmoqda", "type": "status"},
        {"value": Order.ProductionStatus.READY, "icon": "✅", "label": "Tayyor", "type": "status"},
        {"value": Order.ProductionStatus.SHIPPING, "icon": "🚚", "label": "Yetkazilmoqda", "type": "status"},
        {"value": Order.ProductionStatus.DELIVERED, "icon": "🏁", "label": "Yetkazildi", "type": "status"},
        {"type": "divider", "label": "Yakuniy"},
        {"value": "COMPLETED", "icon": "✔️", "label": "Yakunlandi", "type": "virtual"},  # Virtual: DELIVERED + PAID
        {"value": Order.ProductionStatus.CANCELLED, "icon": "❌", "label": "Bekor qilindi", "type": "status"},
    ]


@login_required
def order_list(request):
    from calculator.models import CalculatorSettings
    from django.utils import timezone
    from collections import defaultdict

    # USD kursini olish
    try:
        settings = CalculatorSettings.objects.first()
        usd_rate = int(settings.usd_rate) if settings and settings.usd_rate else 12850
    except:
        usd_rate = 12850

    status_tabs = _order_status_tabs()
    valid_status_values = {item["value"] for item in status_tabs if item.get("type") not in ["divider"]}
    selected_status = request.GET.get("status_tab", "ALL")
    if selected_status not in valid_status_values:
        selected_status = "ALL"

    # Base querysets: if superuser, show all; else only own orders
    can_view_all_orders = _can_user_view_all_orders(request.user)

    if can_view_all_orders:
        orders_qs = Order.objects.all().prefetch_related("items").select_related("sold_by").order_by("-updated_at")
    else:
        orders_qs = Order.objects.filter(sold_by=request.user).prefetch_related("items").select_related("sold_by").order_by("-updated_at")

    # Common filters
    sold_by = request.GET.get("sold_by")
    customer_phone = request.GET.get("customer_phone", "").strip()
    date_from = request.GET.get("date_from")
    date_to = request.GET.get("date_to")

    def apply_filters(qs):
        if sold_by and can_view_all_orders:
            qs = qs.filter(sold_by_id=sold_by)
        if customer_phone:
            # Faqat raqamlarni olish va qidirish
            import re
            digits_only = re.sub(r'\D', '', customer_phone)
            if digits_only:
                # Ikki xil qidirish: to'liq format va qisqa format
                from django.db.models import Q
                # 998 bilan boshlangan raqamni ham, boshlangani yo'qni ham qidirish
                if digits_only.startswith('998') and len(digits_only) > 3:
                    # 998991234567 -> 991234567 ham qidirish
                    short_number = digits_only[3:]
                    qs = qs.filter(
                        Q(customer_phone__icontains=digits_only) | 
                        Q(customer_phone__icontains=short_number)
                    )
                else:
                    # Oddiy qidirish
                    qs = qs.filter(customer_phone__icontains=digits_only)
        if date_from:
            qs = qs.filter(created_at__date__gte=date_from)
        if date_to:
            qs = qs.filter(created_at__date__lte=date_to)
        return qs

    filtered_qs = apply_filters(orders_qs)

    status_counts_map = {}
    for item in status_tabs:
        if item.get("type") == "divider":
            continue
        if item["value"] == "ALL":
            status_counts_map[item["value"]] = filtered_qs.count()
        elif item["value"] == "COMPLETED":  # Virtual tab: DELIVERED + PAID
            status_counts_map[item["value"]] = filtered_qs.filter(
                production_status=Order.ProductionStatus.DELIVERED,
                payment_status=Order.PaymentStatus.PAID
            ).count()
        elif item["value"] == Order.ProductionStatus.DELIVERED:  # Faqat to'lanmagan/qisman yetkazilganlar
            status_counts_map[item["value"]] = filtered_qs.filter(
                production_status=Order.ProductionStatus.DELIVERED
            ).exclude(
                payment_status=Order.PaymentStatus.PAID
            ).count()
        else:
            status_counts_map[item["value"]] = filtered_qs.filter(production_status=item["value"]).count()

    if selected_status == "ALL":
        selected_qs = filtered_qs
    elif selected_status == "COMPLETED":  # Virtual tab: DELIVERED + PAID
        selected_qs = filtered_qs.filter(
            production_status=Order.ProductionStatus.DELIVERED,
            payment_status=Order.PaymentStatus.PAID
        )
    elif selected_status == Order.ProductionStatus.DELIVERED:  # Faqat to'lanmagan/qisman yetkazilganlar
        selected_qs = filtered_qs.filter(
            production_status=Order.ProductionStatus.DELIVERED
        ).exclude(
            payment_status=Order.PaymentStatus.PAID
        )
    else:
        selected_qs = filtered_qs.filter(production_status=selected_status)

    # Group orders by date
    def group_by_date(orders):
        groups = defaultdict(list)
        today = timezone.now().date()
        for order in orders:
            order_date = order.created_at.date()
            groups[order_date].append(order)
        # Sort each group by created_at time (newest first within the day)
        for date in groups:
            groups[date].sort(key=lambda x: x.created_at, reverse=True)
        # Sort groups by date descending
        sorted_groups = sorted(groups.items(), key=lambda x: x[0], reverse=True)
        # Add "is_today" flag
        return [(date, _attach_order_permissions(request.user, orders), date == today) for date, orders in sorted_groups]

    status_groups = group_by_date(selected_qs)

    # For superusers, show all regular sellers in filter; for others, no seller filter
    if can_view_all_orders:
        sellers = User.objects.filter(is_superuser=False).order_by("first_name", "username")
    else:
        sellers = []

    return render(request, "orders/order_list.html", {
        "active_nav": "orders",
        "status_tab": selected_status,
        "selected_status_label": next((item["label"] for item in status_tabs if item.get("value") == selected_status), ""),
        "status_groups": status_groups,
        "status_tabs": [
            {
                **item,
                "count": status_counts_map.get(item.get("value"), 0) if "value" in item else None,
            }
            for item in status_tabs
        ],
        "sellers": sellers,
        "can_view_all_orders": can_view_all_orders,
        "usd_rate": usd_rate,
        "filters": {
            "status_tab": selected_status,
            "sold_by": sold_by,
            "customer_phone": customer_phone,
            "date_from": date_from,
            "date_to": date_to,
            "date_from_display": _format_filter_date(date_from),
            "date_to_display": _format_filter_date(date_to),
        },
        "delivery_choices": Order.DeliveryType.choices,
        "payment_status_choices": Order.PaymentStatus.choices,
        "production_status_choices": Order.production_status_choices_for_delivery(Order.DeliveryType.DELIVERY),
    })


def _form_context(user, order, initial, extra=None):
    """order_form.html uchun umumiy context yasaydi."""
    from calculator.models import CalculatorSettings

    # USD kursini olish
    try:
        settings = CalculatorSettings.objects.first()
        usd_rate = int(settings.usd_rate) if settings and settings.usd_rate else 12850
    except:
        usd_rate = 12850

    src = initial or {}
    o = order  # None yoki Order instance

    # Get values
    v_delivery_type = src.get("delivery_type", getattr(o, "delivery_type", ""))
    v_payment_type = src.get("payment_type", getattr(o, "payment_type", ""))
    v_payment_status = src.get("payment_status", getattr(o, "payment_status", ""))
    v_production_status = src.get("production_status", getattr(o, "production_status", ""))
    
    # Default status for new orders - AGREED
    if not v_production_status and not o:
        v_production_status = Order.ProductionStatus.AGREED

    # Create choice label maps
    delivery_label_map = dict(Order.DeliveryType.choices)
    payment_type_label_map = dict(Order.PaymentType.choices)
    payment_status_label_map = dict(Order.PaymentStatus.choices)
    all_production_status_label_map = dict(
        Order.production_status_choices_for_delivery(v_delivery_type or Order.DeliveryType.DELIVERY)
    )
    production_status_choices = _allowed_production_status_choices_for_user(
        user,
        v_delivery_type or Order.DeliveryType.DELIVERY,
        getattr(o, "production_status", None),
        o,
    )

    raw_delivery_time = src.get("delivery_time", getattr(o, "delivery_time", ""))
    if hasattr(raw_delivery_time, "strftime"):
        formatted_delivery_time = raw_delivery_time.strftime("%H:%M")
    elif isinstance(raw_delivery_time, str):
        formatted_delivery_time = raw_delivery_time[:5] if raw_delivery_time else ""
    else:
        formatted_delivery_time = str(raw_delivery_time)[:5] if raw_delivery_time else ""

    ctx = {
        "v_customer_name":  src.get("customer_name",  getattr(o, "customer_name",  "")),
        "v_customer_phone": src.get("customer_phone", getattr(o, "customer_phone", "")),
        "v_region":         src.get("region",         getattr(o, "region",         "")),
        "v_district":       src.get("district",       getattr(o, "district",       "")),
        "v_city":           src.get("city",            getattr(o, "city",           "")),
        "v_landmark":       src.get("landmark",       getattr(o, "landmark",       "")),
        "v_delivery_type":  v_delivery_type,
        "v_delivery_type_label": delivery_label_map.get(v_delivery_type, ""),
        "v_payment_type":   v_payment_type,
        "v_payment_type_label": payment_type_label_map.get(v_payment_type, ""),
        "v_payment_status": v_payment_status,
        "v_payment_status_label": payment_status_label_map.get(v_payment_status, ""),
        "v_production_status": v_production_status,
        "v_production_status_label": all_production_status_label_map.get(v_production_status, ""),
        "v_delivery_date":  src.get("delivery_date", str(getattr(o, "delivery_date", "")) if getattr(o, "delivery_date", None) else ""),
        "v_delivery_time":  formatted_delivery_time,
        "v_partial_amount": src.get("partial_amount", str(getattr(o, "partial_amount", "0"))),
        "v_total_usd":      src.get("total_price_usd",str(getattr(o, "total_price_usd", "0"))),
        "v_total_uzs":      src.get("total_price_uzs",str(getattr(o, "total_price_uzs", "0"))),
        "v_notes":          src.get("notes",          getattr(o, "notes",          "")),
        "order": o,
        "usd_rate": usd_rate,  # USD kursini qo'shamiz
        "active_nav": "orders",
        "delivery_choices": Order.DeliveryType.choices,
        "payment_type_choices": Order.PaymentType.choices,
        "payment_status_choices": Order.PaymentStatus.choices,
        "production_status_choices": production_status_choices,
        "production_status_groups_json": _production_status_groups_json_for_user(user),
        "all_production_status_groups_json": _all_production_status_groups_json(),
        "can_edit_production_status": _can_user_change_production_status(
            user,
            getattr(o, "production_status", None),
            o,
        ),
    }

    # URL parametrlaridan kelgan items ma'lumotlarini qo'shish
    if "items" in src:
        ctx["initial_items"] = src["items"]

    if extra:
        ctx.update(extra)
    return ctx


@login_required
def order_create(request):
    if request.method == "POST":
        try:
            order = _save_order(request, instance=None)
            if request.headers.get("X-Requested-With") == "XMLHttpRequest":
                return JsonResponse({
                    "ok": True,
                    "order_id": order.pk,
                    "order_number": order.display_order_number,
                    "redirect": f"/orders/{order.pk}/",
                })
            return redirect("orders:detail", pk=order.pk)
        except ValueError as e:
            error = str(e)
            if request.headers.get("X-Requested-With") == "XMLHttpRequest":
                return JsonResponse({"ok": False, "error": error}, status=400)
            
            # Items ma'lumotlarini POST dan olish
            initial_data = dict(request.POST.items())
            config_labels = request.POST.getlist("config_label")
            quantities = request.POST.getlist("quantity")
            unit_prices_uzs = request.POST.getlist("unit_price_uzs")
            
            if config_labels:
                items_data = []
                for i, label in enumerate(config_labels):
                    qty = int(quantities[i]) if i < len(quantities) and quantities[i].isdigit() else 1
                    price_uzs = _parse_decimal(unit_prices_uzs[i] if i < len(unit_prices_uzs) else "0")
                    
                    try:
                        from calculator.models import CalculatorSettings
                        settings = CalculatorSettings.objects.first()
                        usd_rate = settings.usd_rate if settings and settings.usd_rate else Decimal("12850")
                        price_usd = price_uzs / usd_rate if usd_rate > 0 else Decimal("0")
                    except:
                        price_usd = Decimal("0")
                    
                    items_data.append({
                        "config_label": label,
                        "quantity": qty,
                        "unit_price_usd": price_usd,
                        "unit_price_uzs": price_uzs
                    })
                initial_data["items"] = items_data
            
            return render(request, "orders/order_form.html",
                          _form_context(request.user, None, initial_data, {"error": error}))

    # GET request - URL parametrlaridan ma'lumotlarni olish
    initial = {}
    if request.GET:
        config_labels = request.GET.getlist("config_label")
        quantities = request.GET.getlist("quantity")
        unit_prices_uzs = request.GET.getlist("unit_price_uzs")

        if config_labels:
            items_data = []
            for i, label in enumerate(config_labels):
                qty = int(quantities[i]) if i < len(quantities) and quantities[i].isdigit() else 1
                price_uzs = _parse_decimal(unit_prices_uzs[i] if i < len(unit_prices_uzs) else "0")

                # So'mdan USD ga konvertatsiya qilish
                try:
                    from calculator.models import CalculatorSettings
                    settings = CalculatorSettings.objects.first()
                    usd_rate = settings.usd_rate if settings and settings.usd_rate else Decimal("12850")
                    price_usd = price_uzs / usd_rate if usd_rate > 0 else Decimal("0")
                except:
                    price_usd = Decimal("0")

                items_data.append({
                    "config_label": label,
                    "quantity": qty,
                    "unit_price_usd": price_usd,
                    "unit_price_uzs": price_uzs
                })
            initial["items"] = items_data

        total_uzs = request.GET.get("total_price_uzs", "")
        initial["total_price_uzs"] = total_uzs

        # USD ga konvertatsiya
        if total_uzs:
            try:
                from calculator.models import CalculatorSettings
                settings = CalculatorSettings.objects.first()
                usd_rate = settings.usd_rate if settings and settings.usd_rate else Decimal("12850")
                total_usd = _parse_decimal(total_uzs) / usd_rate if usd_rate > 0 else Decimal("0")
                initial["total_price_usd"] = str(total_usd)
            except:
                pass

    return render(request, "orders/order_form.html", _form_context(request.user, None, initial))


@login_required
def order_detail(request, pk):
    from calculator.models import CalculatorSettings

    # USD kursini olish
    try:
        settings = CalculatorSettings.objects.first()
        usd_rate = int(settings.usd_rate) if settings and settings.usd_rate else 12850
    except:
        usd_rate = 12850

    order = get_object_or_404(
        _orders_for_user(request.user).prefetch_related("items").select_related("sold_by"),
        pk=pk,
    )
    if request.method == "GET" and request.GET.get("edit") == "1":
        if not _can_user_edit_order(request.user, order):
            raise PermissionDenied("Sizda buyurtmani tahrirlash huquqi yo'q.")
        return render(request, "orders/order_form.html", _form_context(request.user, order, {}))

    if request.method == "POST":
        if not _can_user_edit_order(request.user, order):
            raise PermissionDenied("Sizda buyurtmani tahrirlash huquqi yo'q.")
        try:
            _save_order(request, instance=order)
            return redirect("orders:list")
        except ValueError as e:
            error = str(e)
            
            # Items ma'lumotlarini POST dan olish
            initial_data = dict(request.POST.items())
            config_labels = request.POST.getlist("config_label")
            quantities = request.POST.getlist("quantity")
            unit_prices_uzs = request.POST.getlist("unit_price_uzs")
            
            if config_labels:
                items_data = []
                for i, label in enumerate(config_labels):
                    qty = int(quantities[i]) if i < len(quantities) and quantities[i].isdigit() else 1
                    price_uzs = _parse_decimal(unit_prices_uzs[i] if i < len(unit_prices_uzs) else "0")
                    
                    try:
                        from calculator.models import CalculatorSettings
                        settings = CalculatorSettings.objects.first()
                        usd_rate = settings.usd_rate if settings and settings.usd_rate else Decimal("12850")
                        price_usd = price_uzs / usd_rate if usd_rate > 0 else Decimal("0")
                    except:
                        price_usd = Decimal("0")
                    
                    items_data.append({
                        "config_label": label,
                        "quantity": qty,
                        "unit_price_usd": price_usd,
                        "unit_price_uzs": price_uzs
                    })
                initial_data["items"] = items_data
            
            return render(request, "orders/order_form.html",
                          _form_context(request.user, order, initial_data, {"error": error}))

    allowed_production_status_choices = _allowed_production_status_choices_for_user(
        request.user,
        order.delivery_type,
        order.production_status,
        order,
    )
    allowed_production_status_values = {value for value, _ in allowed_production_status_choices}

    return render(request, "orders/order_detail.html", {
        "active_nav": "orders",
        "order": order,
        "usd_rate": usd_rate,
        "status_flow": Order.production_status_flow(order.delivery_type),
        "current_status_index": (
            Order.production_status_flow(order.delivery_type).index(order.production_status)
            if order.production_status in Order.production_status_flow(order.delivery_type)
            else -1
        ),
        "delivery_choices": Order.DeliveryType.choices,
        "payment_type_choices": Order.PaymentType.choices,
        "payment_status_choices": Order.PaymentStatus.choices,
        "production_status_choices": allowed_production_status_choices,
        "can_edit_order": _can_user_edit_order(request.user, order),
        "can_delete_order": _can_user_delete_order(request.user),
        "can_cancel_order": _can_user_cancel_order(request.user, order),
        "can_change_production_status": _can_user_change_production_status(
            request.user,
            order.production_status,
            order,
        ),
        "show_current_status_option": order.production_status not in allowed_production_status_values,
        "current_status_label": _all_production_status_choices_map(order.delivery_type).get(
            order.production_status,
            order.production_status,
        ),
    })


@login_required
@require_POST
def order_update_status(request, pk):
    order = get_object_or_404(_orders_for_user(request.user), pk=pk)
    new_status = request.POST.get("status")
    valid = [s for s, _ in Order.production_status_choices_for_delivery(order.delivery_type)]
    if new_status not in valid:
        return JsonResponse({"error": "Noto'g'ri status."}, status=400)
    if not _can_user_set_production_status(request.user, new_status, order.production_status, order):
        return JsonResponse(
            {"error": "Siz bu buyurtma uchun ushbu statusni o'zgartira olmaysiz."},
            status=403,
        )
    
    # Validatsiya (barcha foydalanuvchilar uchun)
    is_valid, error_msg = _validate_status_transition(order, new_status)
    if not is_valid:
        return JsonResponse({"error": error_msg}, status=400)
    
    order.production_status = new_status
    # delivered_at ni DELIVERED statusiga o'tganda to'ldirish
    if new_status == Order.ProductionStatus.DELIVERED:
        if not order.delivered_at:
            order.delivered_at = timezone.now()
    order.save(update_fields=["production_status", "delivered_at", "updated_at"])
    response_data = {
        "status": order.production_status,
        "status_label": Order.production_status_label(order.production_status, order.delivery_type),
    }
    if request.headers.get("X-Requested-With") == "XMLHttpRequest":
        return JsonResponse(response_data)
    return redirect("orders:detail", pk=order.pk)


@login_required
def order_delete(request, pk):
    if not _can_user_delete_order(request.user):
        raise PermissionDenied("Sizda buyurtmani o'chirish huquqi yo'q.")

    order = get_object_or_404(_orders_for_user(request.user), pk=pk)
    if request.method == "POST":
        order.delete()
        return redirect("orders:list")
    return redirect("orders:detail", pk=pk)


@login_required
def export_orders_excel(request):
    """
    COMPLETED statusidagi orderlarni Excel ga export qilish
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, numbers
    from django.http import HttpResponse
    from datetime import datetime
    import re
    
    def format_phone_number(phone):
        """
        Telefon raqamni +998 (XX) XXX-XX-XX formatiga o'zgartirish
        Input: 943456789 yoki +998943456789 yoki boshqa format
        Output: +998 (94) 345-67-89
        """
        if not phone:
            return ''
        
        # Faqat raqamlarni olish
        digits = re.sub(r'\D', '', str(phone))
        
        # Agar 998 bilan boshlanmasa, uni qo'shish
        if not digits.startswith('998'):
            digits = '998' + digits
        
        # 998 dan keyingi raqamlar
        if len(digits) >= 12:
            # Format: +998 (XX) XXX-XX-XX
            code = digits[3:5]
            part1 = digits[5:8]
            part2 = digits[8:10]
            part3 = digits[10:12]
            return f'+998 ({code}) {part1}-{part2}-{part3}'
        
        # Agar format noto'g'ri bo'lsa, asl qiymatni qaytarish
        return phone
    
    # Sanalarni olish
    date_from_str = request.GET.get('export_date_from', '')
    date_to_str = request.GET.get('export_date_to', '')
    
    if not date_from_str or not date_to_str:
        return JsonResponse({'error': 'Sanalar kiritilmagan'}, status=400)
    
    try:
        date_from = datetime.strptime(date_from_str, '%Y-%m-%d').date()
        date_to = datetime.strptime(date_to_str, '%Y-%m-%d').date()
    except ValueError:
        return JsonResponse({'error': 'Noto\'g\'ri sana formati'}, status=400)
    
    # Orderlarni filter qilish
    orders_qs = Order.objects.filter(
        production_status=Order.ProductionStatus.DELIVERED,
        payment_status=Order.PaymentStatus.PAID,
        created_at__date__gte=date_from,
        created_at__date__lte=date_to
    ).select_related('sold_by').prefetch_related('items').order_by('created_at')
    
    # Faqat o'z orderlarini export qilish (admin bo'lmasa)
    if not (request.user.is_superuser or request.user.has_perm('orders.view_order')):
        orders_qs = orders_qs.filter(sold_by=request.user)
    
    # Excel yaratish
    wb = Workbook()
    ws = wb.active
    ws.title = "Yakunlangan Orderlar"
    
    # Header styling
    header_fill = PatternFill(start_color='0F766E', end_color='0F766E', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    # Header qatorlari
    headers = [
        'Order #',
        'Sana',
        'Mijoz',
        'Telefon',
        'Manzil',
        'Yetkazish turi',
        'To\'lov turi',
        'To\'lov holati',
        'Konfiguratsiya',
        'Soni',
        'Narxi (so\'m)',
        'Jami summa (so\'m)',
        'Izoh',
        'Sotuvchi',
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Ma'lumotlarni yozish
    row_num = 2
    for order in orders_qs:
        # Manzilni birlashtirish
        address_parts = []
        if order.region:
            address_parts.append(order.region)
        if order.district:
            address_parts.append(order.district)
        if order.city:
            address_parts.append(order.city)
        if order.landmark:
            address_parts.append(order.landmark)
        full_address = ', '.join(address_parts)
        
        # Telefon raqamni formatlash
        formatted_phone = format_phone_number(order.customer_phone)
        
        # Orderning barcha itemlarini olish
        items = list(order.items.all())
        
        if items:
            # Orderning boshlanish qatori
            start_row = row_num
            
            # Birinchi item uchun to'liq ma'lumot
            first_item = items[0]
            ws.cell(row=row_num, column=1, value=order.display_order_number)
            ws.cell(row=row_num, column=2, value=order.created_at.strftime('%d.%m.%Y %H:%M'))
            ws.cell(row=row_num, column=3, value=order.customer_name)
            ws.cell(row=row_num, column=4, value=formatted_phone)
            ws.cell(row=row_num, column=5, value=full_address)
            ws.cell(row=row_num, column=6, value=order.get_delivery_type_display())
            ws.cell(row=row_num, column=7, value=order.get_payment_type_display())
            ws.cell(row=row_num, column=8, value=order.get_payment_status_display())
            ws.cell(row=row_num, column=9, value=first_item.config_label)
            ws.cell(row=row_num, column=10, value=first_item.quantity)
            
            # Narxi - number format bilan
            price_cell = ws.cell(row=row_num, column=11, value=float(first_item.unit_price_uzs))
            price_cell.number_format = '#,##0'
            
            # Jami summa - number format bilan
            total_cell = ws.cell(row=row_num, column=12, value=float(order.total_price_uzs))
            total_cell.number_format = '#,##0'
            
            ws.cell(row=row_num, column=13, value=order.notes or '')
            ws.cell(row=row_num, column=14, value=order.sold_by.get_full_name() or order.sold_by.username)
            row_num += 1
            
            # Qolgan itemlar uchun faqat konfiguratsiya, soni va narxi
            for item in items[1:]:
                ws.cell(row=row_num, column=9, value=item.config_label)
                ws.cell(row=row_num, column=10, value=item.quantity)
                
                # Narxi - number format bilan
                item_price_cell = ws.cell(row=row_num, column=11, value=float(item.unit_price_uzs))
                item_price_cell.number_format = '#,##0'
                
                row_num += 1
            
            # Agar bir nechta item bo'lsa, umumiy ustunlarni merge qilish
            end_row = row_num - 1
            if end_row > start_row:
                # Order #, Sana, Mijoz, Telefon, Manzil, Yetkazish, To'lov turi, To'lov holati, Jami summa, Izoh, Sotuvchi
                ws.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)  # Order #
                ws.merge_cells(start_row=start_row, start_column=2, end_row=end_row, end_column=2)  # Sana
                ws.merge_cells(start_row=start_row, start_column=3, end_row=end_row, end_column=3)  # Mijoz
                ws.merge_cells(start_row=start_row, start_column=4, end_row=end_row, end_column=4)  # Telefon
                ws.merge_cells(start_row=start_row, start_column=5, end_row=end_row, end_column=5)  # Manzil
                ws.merge_cells(start_row=start_row, start_column=6, end_row=end_row, end_column=6)  # Yetkazish turi
                ws.merge_cells(start_row=start_row, start_column=7, end_row=end_row, end_column=7)  # To'lov turi
                ws.merge_cells(start_row=start_row, start_column=8, end_row=end_row, end_column=8)  # To'lov holati
                ws.merge_cells(start_row=start_row, start_column=12, end_row=end_row, end_column=12) # Jami summa
                ws.merge_cells(start_row=start_row, start_column=13, end_row=end_row, end_column=13) # Izoh
                ws.merge_cells(start_row=start_row, start_column=14, end_row=end_row, end_column=14) # Sotuvchi
                
                # Merge qilingan celllarni chapdan tekislash va vertikal top
                for col in [1, 2, 3, 4, 5, 6, 7, 8, 12, 13, 14]:
                    cell = ws.cell(row=start_row, column=col)
                    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            
            # Merge qilinmagan celllarni ham chapdan tekislash
            for r in range(start_row, end_row + 1):
                for col in [9, 10, 11]:
                    cell = ws.cell(row=r, column=col)
                    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        else:
            # Agar itemlar bo'lmasa (bo'lmasligi kerak, lekin xavfsizlik uchun)
            ws.cell(row=row_num, column=1, value=order.display_order_number)
            ws.cell(row=row_num, column=2, value=order.created_at.strftime('%d.%m.%Y %H:%M'))
            ws.cell(row=row_num, column=3, value=order.customer_name)
            ws.cell(row=row_num, column=4, value=formatted_phone)
            ws.cell(row=row_num, column=5, value=full_address)
            ws.cell(row=row_num, column=6, value=order.get_delivery_type_display())
            ws.cell(row=row_num, column=7, value=order.get_payment_type_display())
            ws.cell(row=row_num, column=8, value=order.get_payment_status_display())
            ws.cell(row=row_num, column=9, value='')
            ws.cell(row=row_num, column=10, value='')
            ws.cell(row=row_num, column=11, value='')
            
            # Jami summa - number format bilan
            total_cell = ws.cell(row=row_num, column=12, value=float(order.total_price_uzs))
            total_cell.number_format = '#,##0'
            
            ws.cell(row=row_num, column=13, value=order.notes or '')
            ws.cell(row=row_num, column=14, value=order.sold_by.get_full_name() or order.sold_by.username)
            
            # Alignment qo'shish
            for col in range(1, 15):
                cell = ws.cell(row=row_num, column=col)
                cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            
            row_num += 1
    
    # Column widthlarni sozlash
    column_widths = {
        'A': 15,  # Order #
        'B': 18,  # Sana
        'C': 20,  # Mijoz
        'D': 20,  # Telefon
        'E': 45,  # Manzil (to'liq)
        'F': 18,  # Yetkazish turi
        'G': 15,  # To'lov turi
        'H': 18,  # To'lov holati
        'I': 50,  # Konfiguratsiya
        'J': 10,  # Soni
        'K': 15,  # Narxi
        'L': 18,  # Jami summa
        'M': 30,  # Izoh
        'N': 20,  # Sotuvchi
    }
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # Response yaratish
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'yakunlangan_orderlar_{date_from_str}_{date_to_str}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response


@login_required
def leaderboard(request):
    from django.db.models import Sum, Count
    from django.utils import timezone
    from datetime import timedelta

    period = request.GET.get("period", "all")
    now = timezone.now()

    if period == "today":
        date_filter = Q(created_at__date=now.date())
    elif period == "week":
        week_start = now.date() - timezone.timedelta(days=now.weekday())
        date_filter = Q(created_at__date__gte=week_start)
    elif period == "month":
        date_filter = Q(created_at__year=now.year, created_at__month=now.month)
    else:
        date_filter = Q()

    active_filter = date_filter & ~Q(production_status=Order.ProductionStatus.CANCELLED)
    cancelled_filter = date_filter & Q(production_status=Order.ProductionStatus.CANCELLED)

    sellers = User.objects.filter(
        is_superuser=False,
        orders__isnull=False
    ).distinct()

    board = []
    for seller in sellers:
        agg = Order.objects.filter(active_filter, sold_by=seller).aggregate(
            total_usd=Sum("total_price_usd"),
            total_uzs=Sum("total_price_uzs"),
            order_count=Count("id"),
        )
        units = OrderItem.objects.filter(
            order__sold_by=seller,
            order__in=Order.objects.filter(active_filter, sold_by=seller),
        ).aggregate(total_units=Sum("quantity"))["total_units"] or 0

        cancelled = Order.objects.filter(cancelled_filter, sold_by=seller).count()
        
        # To'langan summalarni hisoblash (yakunlanmagan orderlar uchun ham)
        seller_orders = Order.objects.filter(active_filter, sold_by=seller)
        paid_amount_uzs = Decimal("0")
        paid_amount_usd = Decimal("0")
        
        for order in seller_orders:
            if order.payment_status == Order.PaymentStatus.PAID:
                paid_amount_uzs += order.total_price_uzs
                paid_amount_usd += order.total_price_usd
            elif order.payment_status == Order.PaymentStatus.PARTIAL:
                paid_amount_uzs += order.partial_amount
                # USD ga konvertatsiya
                try:
                    from calculator.models import CalculatorSettings
                    settings = CalculatorSettings.objects.first()
                    usd_rate = settings.usd_rate if settings and settings.usd_rate else Decimal("12850")
                    paid_amount_usd += order.partial_amount / usd_rate if usd_rate > 0 else Decimal("0")
                except:
                    pass

        board.append({
            "seller": seller,
            "total_usd": agg["total_usd"] or Decimal("0"),
            "total_uzs": agg["total_uzs"] or Decimal("0"),
            "paid_amount_uzs": paid_amount_uzs,
            "paid_amount_usd": paid_amount_usd,
            "unpaid_amount_uzs": (agg["total_uzs"] or Decimal("0")) - paid_amount_uzs,
            "order_count": agg["order_count"] or 0,
            "units_sold": units,
            "cancelled": cancelled,
        })

    board.sort(key=lambda x: (x["units_sold"], x["total_usd"]), reverse=True)

    return render(request, "orders/leaderboard.html", {
        "active_nav": "orders",
        "board": board,
        "period": period,
    })


@login_required
def seller_profile(request):
    profile_url = redirect("calculator:profile")
    query_string = request.GET.urlencode()
    if query_string:
        profile_url["Location"] = f"{profile_url['Location']}?{query_string}"
    return profile_url


def _save_order(request, instance=None):
    """Create or update an Order + its OrderItems from POST data."""
    post = request.POST

    customer_name = post.get("customer_name", "").strip()
    customer_phone = post.get("customer_phone", "").strip()
    region = post.get("region", "").strip()
    district = post.get("district", "").strip()
    city = post.get("city", "").strip()
    landmark = post.get("landmark", "").strip()
    delivery_type = post.get("delivery_type", "").strip()
    payment_type = post.get("payment_type", "").strip()
    payment_status = post.get("payment_status", "").strip()
    production_status = post.get("production_status", "").strip()
    if not production_status:
        production_status = getattr(instance, "production_status", "") or Order.ProductionStatus.AGREED

    delivery_date_str = post.get("delivery_date", "").strip()
    delivery_time = post.get("delivery_time", "").strip()
    partial_amount_str = post.get("partial_amount", "0").strip()

    notes = post.get("notes", "").strip()

    errors = []

    delivery_date = None
    if delivery_date_str:
        from datetime import datetime
        try:
            delivery_date = datetime.strptime(delivery_date_str, "%Y-%m-%d").date()
        except ValueError:
            pass

    clean_partial = "".join(c for c in partial_amount_str if c.isdigit() or c in ".,")
    partial_amount = _parse_decimal(clean_partial, Decimal("0"))

    if not customer_name:
        errors.append("Mijoz ismi kiritilmadi.")
    if not customer_phone:
        errors.append("Telefon raqami kiritilmadi.")
    if not region:
        errors.append("Viloyat kiritilmadi.")
    if not district:
        errors.append("Tuman kiritilmadi.")
    if delivery_type not in [d for d, _ in Order.DeliveryType.choices]:
        errors.append("Yetkazish turi tanlanmadi.")
    if payment_type not in [p for p, _ in Order.PaymentType.choices]:
        errors.append("To'lov turi tanlanmadi.")
    if payment_status not in [p for p, _ in Order.PaymentStatus.choices]:
        errors.append("To'lov statusi tanlanmadi.")
    valid_production_statuses = [s for s, _ in Order.production_status_choices_for_delivery(delivery_type)]
    if production_status not in valid_production_statuses:
        errors.append("Status tanlovi yetkazish turiga mos emas.")
    if not _can_user_set_production_status(
        request.user,
        production_status,
        getattr(instance, "production_status", None),
        instance,
    ):
        errors.append("Siz bu buyurtma uchun ushbu statusni o'zgartira olmaysiz.")
    
    # Status validatsiyasi (yangi order va edit qilish paytida)
    # Temporary order object for validation
    if instance:
        temp_order = instance
        current_prod_status = instance.production_status
    else:
        # Yangi order yaratilayotgan - default status AGREED
        temp_order = type('obj', (object,), {
            'production_status': Order.ProductionStatus.AGREED,
            'payment_status': payment_status,
            'delivery_type': delivery_type,
        })()
        current_prod_status = Order.ProductionStatus.AGREED
    
    # Payment status ni yangilaymiz
    temp_order.payment_status = payment_status
    temp_order.delivery_type = delivery_type
    
    # Validatsiya chaqirish
    is_valid, error_msg = _validate_status_transition(temp_order, production_status)
    if not is_valid:
        errors.append(error_msg)
    
    if delivery_type and (not delivery_date or not delivery_time):
        errors.append("Yetkazib berish sanasi va vaqti kiritilmadi.")

    config_labels = post.getlist("config_label")
    quantities = post.getlist("quantity")
    unit_prices_uzs = post.getlist("unit_price_uzs")

    items_data = []
    total_price_uzs = Decimal("0")
    # Get USD rate once for all calculations
    try:
        from calculator.models import CalculatorSettings
        settings = CalculatorSettings.objects.first()
        usd_rate = settings.usd_rate if settings and settings.usd_rate else Decimal("12850")
    except:
        usd_rate = Decimal("12850")

    for i, label in enumerate(config_labels):
        label = label.strip()
        if not label:
            continue
        qty = int(quantities[i]) if i < len(quantities) and quantities[i].isdigit() else 1
        price_uzs = _parse_decimal(unit_prices_uzs[i] if i < len(unit_prices_uzs) else "0")

        # So'mdan USD ga konvertatsiya qilish (database uchun)
        price_usd = price_uzs / usd_rate if usd_rate > 0 else Decimal("0")

        items_data.append({"config_label": label, "quantity": qty, "unit_price_usd": price_usd, "unit_price_uzs": price_uzs})
        total_price_uzs += price_uzs * qty

    total_price_usd = total_price_uzs / usd_rate if usd_rate > 0 else Decimal("0")

    if not items_data:
        errors.append("Kamida bitta mahsulot qo'shing.")

    if errors:
        raise ValueError(" | ".join(errors))

    if instance is None:
        order = Order(sold_by=request.user)
    else:
        order = instance

    order.customer_name = customer_name
    order.customer_phone = customer_phone
    order.region = region
    order.district = district
    order.city = city
    order.landmark = landmark
    order.delivery_type = delivery_type
    order.payment_type = payment_type
    order.payment_status = payment_status
    order.production_status = production_status
    order.total_price_usd = total_price_usd
    order.total_price_uzs = total_price_uzs
    order.delivery_date = delivery_date
    order.delivery_time = delivery_time
    order.partial_amount = partial_amount if payment_status == Order.PaymentStatus.PARTIAL else Decimal("0")
    order.notes = notes

    if order.production_status == Order.ProductionStatus.DELIVERED and not order.delivered_at:
        order.delivered_at = timezone.now()

    order.save()

    if not order.order_number:
        order.order_number = f"ORDER#{order.pk}"
        order.save(update_fields=["order_number"])

    # Replace items
    order.items.all().delete()
    for item in items_data:
        OrderItem.objects.create(order=order, **item)

    return order
