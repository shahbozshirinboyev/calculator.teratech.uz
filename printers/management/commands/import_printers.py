from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path

import openpyxl
from django.core.management.base import BaseCommand

from printers.models import Printer


class Command(BaseCommand):
    help = "Import printers from an Excel file. Each row: Name | Price"

    def add_arguments(self, parser):
        parser.add_argument(
            "path",
            nargs="?",
            default="data/Data-20.06.2026.xlsx",
            help="Excel file path.",
        )
        parser.add_argument(
            "--sheet",
            default="Printer's",
            help="Sheet name to import from.",
        )

    def handle(self, *args, **options):
        workbook_path = Path(options["path"])
        if not workbook_path.is_absolute():
            workbook_path = Path.cwd() / workbook_path
        if not workbook_path.exists():
            self.stderr.write(self.style.ERROR(f"File not found: {workbook_path}"))
            return

        wb = openpyxl.load_workbook(workbook_path, data_only=True, read_only=True)
        sheet_name = options["sheet"]

        if sheet_name not in wb.sheetnames:
            self.stderr.write(self.style.ERROR(
                f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}"
            ))
            return

        sheet = wb[sheet_name]
        imported = 0

        for row in range(2, sheet.max_row + 1):
            name = sheet.cell(row, 1).value
            price = sheet.cell(row, 3).value  # col 3 = narx (col 2 = model kodi)
            if name is None or price is None:
                continue
            name = str(name).strip()
            if not name:
                continue
            try:
                price = Decimal(str(price)).quantize(
                    Decimal("0.01"), rounding=ROUND_HALF_UP
                )
            except Exception:
                continue

            Printer.objects.update_or_create(
                name=name,
                defaults={"price": price, "is_active": True},
            )
            imported += 1

        self.stdout.write(self.style.SUCCESS(
            f"Imported {imported} printers from '{sheet_name}'"
        ))
