from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path

import openpyxl
from django.core.management.base import BaseCommand

from laptops.models import Laptop
from monitors.models import Monitor
from printers.models import Printer


SHEET_MODEL_MAP = {
    "Monitor's": Monitor,
    "Monitors": Monitor,
    "Printer's": Printer,
    "Printers": Printer,
    "Laptop's": Laptop,
    "Laptops": Laptop,
    "Laptop": Laptop,
}


class Command(BaseCommand):
    help = (
        "Import monitors, printers, and laptops from a single Excel file. "
        "Detects sheets by name (Monitor's, Printer's, Laptop's)."
    )

    def add_arguments(self, parser):
        parser.add_argument(
            "path",
            nargs="?",
            default="data/Data-20.06.2026.xlsx",
            help="Excel file path. Default: data/Data-20.06.2026.xlsx",
        )

    def handle(self, *args, **options):
        workbook_path = Path(options["path"])
        if not workbook_path.is_absolute():
            workbook_path = Path.cwd() / workbook_path
        if not workbook_path.exists():
            self.stderr.write(self.style.ERROR(f"File not found: {workbook_path}"))
            return

        wb = openpyxl.load_workbook(workbook_path, data_only=True, read_only=True)
        total = 0

        for sheet_name in wb.sheetnames:
            model = SHEET_MODEL_MAP.get(sheet_name)
            if model is None:
                self.stdout.write(f"Skipping unknown sheet: '{sheet_name}'")
                continue

            sheet = wb[sheet_name]
            imported = self._import_sheet(sheet, model)
            total += imported
            self.stdout.write(self.style.SUCCESS(
                f"  {model.__name__}: {imported} items from '{sheet_name}'"
            ))

        self.stdout.write(self.style.SUCCESS(
            f"\nTotal imported: {total} items from {workbook_path.name}"
        ))

    def _import_sheet(self, sheet, model):
        imported = 0
        for row in range(2, sheet.max_row + 1):
            name = sheet.cell(row, 1).value
            # Try col 3 first (Excel format: name | model_code | price), fallback to col 2
            price = sheet.cell(row, 3).value
            if price is None:
                price = sheet.cell(row, 2).value
            if name is None or price is None:
                continue
            name = str(name).strip()
            if not name:
                continue
            try:
                price = Decimal(str(price)).quantize(
                    Decimal("0.01"), rounding=ROUND_HALF_UP
                )
            except Exception:
                continue

            model.objects.update_or_create(
                name=name,
                defaults={"price": price, "is_active": True},
            )
            imported += 1
        return imported
